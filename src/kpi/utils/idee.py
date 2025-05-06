import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference


def _sanitize_sheet_name(name: str) -> str:
    """Sanitize sheet names to be valid in Excel."""
    invalid_chars = r"[:\\/*?[\]]"
    return re.sub(invalid_chars, "", name)[:31]


def in_range(date, start_date, end_date):
    return start_date <= date <= end_date


def _apply_standard_sheet_styling(sheet: Worksheet, headers: list[str]) -> None:
    """Apply consistent styling to a worksheet."""
    # Set column widths
    for i, _ in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = 25

    # Style header row
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        cell.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

    # Freeze header row
    sheet.freeze_panes = "A2"

    # Add auto-filter
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"


def _kunden_pro_jahr(kpi_data: dict[str, list], wb: Workbook, start_date, end_date):
    """Create 'Customers per Year' sheet with enhanced styling."""
    if kpi_data.get("customer_kpis"):
        sheet = wb.create_sheet("Kunden pro Jahr")
        headers = ["Jahr", "Anzahl"]
        sheet.append(headers)

        # Process data
        stats = defaultdict(int)
        for c in kpi_data["customer_kpis"]:
            if in_range(c.registered_at, start_date, end_date):
                stats[c.registered_at.year] += 1

        # Add data rows
        for year, count in sorted(stats.items()):
            sheet.append([year, count])

        # Apply standard styling
        _apply_standard_sheet_styling(sheet, headers)

        # Create chart
        chart = BarChart()
        chart.title = "Kunden pro Jahr"
        chart.style = 13
        chart.y_axis.title = "Anzahl"
        chart.x_axis.title = "Jahr"
        chart.add_data(
            Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
        )
        sheet.add_chart(chart, "D2")


def _transaktion_pro_monat(kpi_data, wb, start_date, end_date):
    """Create 'Average Transaction Amount per Month' sheet."""
    if kpi_data.get("transaction_kpis"):
        sheet = wb.create_sheet("Ø Transaktionshöhe")
        headers = ["Monat", "Ø Betrag"]
        sheet.append(headers)

        # Process data
        monthly = defaultdict(list)
        for t in kpi_data["transaction_kpis"]:
            if in_range(t.created_at, start_date, end_date):
                key = t.created_at.strftime("%Y-%m")
                monthly[key].append(t.amount)

        # Add data rows
        for k, lst in sorted(monthly.items()):
            avg = sum(lst) / len(lst)
            sheet.append([k, avg])

        # Apply standard styling
        _apply_standard_sheet_styling(sheet, headers)

        # Create chart
        chart = LineChart()
        chart.title = "Durchschnittliche Transaktionen je Monat"
        chart.style = 13
        chart.y_axis.title = "Betrag (€)"
        chart.x_axis.title = "Monat"
        chart.add_data(
            Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
        )
        sheet.add_chart(chart, "D2")


def _bewegung(kpi_data, wb, start_date, end_date):
    """Create 'Product Movement' sheet with pie chart."""
    if kpi_data.get("product_movement_kpis"):
        sheet = wb.create_sheet("Bewegung (Kreis)")
        headers = ["Typ", "Anzahl"]
        sheet.append(headers)

        # Process data
        stat = defaultdict(int)
        for m in kpi_data["product_movement_kpis"]:
            if in_range(m.created_at, start_date, end_date):
                stat[m.movement_type] += 1

        # Add data rows
        for t, c in stat.items():
            sheet.append([t, c])

        # Apply standard styling
        _apply_standard_sheet_styling(sheet, headers)

        # Create chart
        chart = PieChart()
        chart.title = "Verkäufe vs. Käufe"
        chart.add_data(
            Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
        )
        sheet.add_chart(chart, "D2")


def _top_5_kunden(kpi_data, wb, start_date, end_date):
    """Create 'Top 5 Customers by Revenue' sheet."""
    if kpi_data.get("order_kpis"):
        sheet = wb.create_sheet("Top Kunden")
        headers = ["Kunde", "Umsatz (€)"]
        sheet.append(headers)

        # Process data
        totals = defaultdict(float)
        for o in kpi_data["order_kpis"]:
            if in_range(o.created_at, start_date, end_date):
                totals[o.user_id] += o.total_price

        # Add data rows
        for uid, val in sorted(totals.items(), key=lambda x: -x[1])[:5]:
            sheet.append([uid, val])

        # Apply standard styling
        _apply_standard_sheet_styling(sheet, headers)

        # Create chart
        chart = BarChart()
        chart.title = "Top 5 Kunden nach Umsatz"
        chart.style = 13
        chart.y_axis.title = "Umsatz (€)"
        chart.add_data(
            Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
        )
        sheet.add_chart(chart, "D2")


def _registrierungen_pro_quartal(kpi_data, wb, start_date, end_date):
    """Create 'Registrations per Quarter' sheet."""
    if kpi_data.get("customer_kpis"):
        sheet = wb.create_sheet("Registrierungen")
        headers = ["Quartal", "Anzahl"]
        sheet.append(headers)

        # Process data
        qstats = defaultdict(int)
        for c in kpi_data["customer_kpis"]:
            if in_range(c.registered_at, start_date, end_date):
                q = f"Q{((c.registered_at.month-1)//3)+1} {c.registered_at.year}"
                qstats[q] += 1

        # Add data rows
        for q, val in sorted(qstats.items()):
            sheet.append([q, val])

        # Apply standard styling
        _apply_standard_sheet_styling(sheet, headers)

        # Create chart
        chart = BarChart()
        chart.title = "Registrierungen pro Quartal"
        chart.style = 13
        chart.y_axis.title = "Anzahl"
        chart.x_axis.title = "Quartal"
        chart.add_data(
            Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
        )
        sheet.add_chart(chart, "D2")


def create_chart(
    chart_type,
    title,
    y_axis_title,
    x_axis_title,
    sheet,
    data_col,
    category_col,
    position,
):
    chart = chart_type()
    chart.title = title
    chart.style = 13
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    chart.add_data(
        Reference(sheet, min_col=data_col, min_row=1, max_row=sheet.max_row),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(sheet, min_col=category_col, min_row=2, max_row=sheet.max_row)
    )
    sheet.add_chart(chart, position)

create_chart(BarChart, "Kunden pro Jahr", "Anzahl", "Jahr", sheet, 2, 1, "D2")

sheet = wb.create_sheet(_sanitize_sheet_name("Kunden pro Jahr"))

for year, count in sorted(stats.items()):
    sheet.append([year, count])
