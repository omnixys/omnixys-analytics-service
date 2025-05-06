# src/kpi/utils/export_utils.py
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Union
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import csv
from loguru import logger
from typing import Any

from pydantic import BaseModel

import re
import uuid
from pathlib import Path
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.chart import PieChart, BarChart, Reference
from loguru import logger


def export_to_csv(kpi_data: dict[str, list], path: Path) -> None:
    import csv

    with open(path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")

        for sheet_name, entries in kpi_data.items():
            if not entries:
                continue

            # 💡 Dynamisch alle Spaltennamen aus dem SQLAlchemy-Modell
            model = entries[0].__class__
            columns = [c.name for c in model.__table__.columns]

            # ✍️ Optional: Abschnittsüberschrift in der CSV
            writer.writerow([f"=== {sheet_name.upper()} ==="])
            writer.writerow(columns)

            for entry in entries:
                writer.writerow([getattr(entry, col) for col in columns])


def _sanitize_sheet_name(name: str) -> str:
    invalid_chars = r"[:\\/*?[\]]"
    return re.sub(invalid_chars, "", name)[:31]


def export_to_excel(kpi_data: dict[str, list], export_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)  # Standard-Sheet löschen

    for sheet_name, entries in kpi_data.items():
        sheet = workbook.create_sheet(_sanitize_sheet_name(sheet_name))
        if not entries:
            sheet.append(["Keine Daten vorhanden"])
            continue

        # Dynamisch Spaltennamen bestimmen
        first = entries[0]
        headers = [c.name for c in first.__table__.columns]
        sheet.append(headers)

        # Werte einfügen
        for entry in entries:
            row = []
            for col in headers:
                value = getattr(entry, col)
                if isinstance(value, (uuid.UUID, datetime)):
                    value = str(value)
                row.append(value)
            sheet.append(row)

    # Logo in erstes Sheet (falls vorhanden)
    try:
        logo_path = Path(__file__).parent.parent / "static/logo.png"
        if logo_path.exists():
            sheet = workbook.worksheets[0]
            logo = ExcelImage(str(logo_path))
            logo.width = 300
            logo.height = 80
            sheet.merge_cells("A1:C4")
            sheet.add_image(logo, "A1")
            logger.debug("Logo eingebettet: %s", logo_path)
    except Exception as e:
        logger.warning("Fehler beim Einfügen des Logos: {}", e)

    # Beispiel-Diagramm: Kunden pro Jahr
    if "customer_kpis" in kpi_data and kpi_data["customer_kpis"]:
        sheet = workbook.create_sheet("Kunden pro Jahr")
        sheet.append(["Jahr", "Anzahl"])
        stats = defaultdict(int)
        for c in kpi_data["customer_kpis"]:
            jahr = c.registered_at.year
            stats[jahr] += 1
        for year, count in sorted(stats.items()):
            sheet.append([year, count])
        chart = BarChart()
        chart.title = "Kunden pro Jahr"
        chart.x_axis.title = "Jahr"
        chart.y_axis.title = "Anzahl"
        data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        sheet.add_chart(chart, "D2")

    # Speichern
    workbook.save(export_path)
    logger.success("Excel erfolgreich exportiert: {}", export_path)
