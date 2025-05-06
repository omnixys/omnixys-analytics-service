from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
import re
import uuid
from loguru import logger


def export_to_excel(
    kpi_data: dict[str, list], export_path: Path, start_date=None, end_date=None
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    def in_range(obj_date):
        if not start_date and not end_date:
            return True
        if start_date and obj_date < start_date:
            return False
        if end_date and obj_date > end_date:
            return False
        return True

    def sanitize_sheet_name(name: str) -> str:
        invalid_chars = r"[:\\/*?[\]]"
        return re.sub(invalid_chars, "", name)[:31]

    def table_border(
        row_idx: int,
        col_idx: int,
        start_row: int,
        start_col: int,
        rows: list[list],
        header: list[str],
    ) -> Border:
        thick = Side(border_style="thick", color="000000")
        thin = Side(border_style="thin", color="000000")
        is_top = row_idx == start_row
        is_bottom = row_idx == start_row + len(rows)
        is_left = col_idx == start_col
        is_right = col_idx == start_col + len(header) - 1
        return Border(
            top=thick if is_top else thin,
            bottom=thick if is_bottom else thin,
            left=thick if is_left else thin,
            right=thick if is_right else thin,
        )

    def add_hyperlinked_sheet_overview(workbook: Workbook):
        sheet = workbook.create_sheet("Übersicht", 0)
        sheet.append(["Sheet", "Beschreibung"])
        for ws in workbook.worksheets[1:]:
            sheet.append([ws.title, f"Analyse von {ws.title.lower()}"])
            cell = sheet.cell(row=sheet.max_row, column=1)
            cell.hyperlink = f"#{ws.title}!A1"
            cell.font = Font(color="0000FF", underline="single")

    for sheet_name, entries in kpi_data.items():
        sheet = wb.create_sheet(sanitize_sheet_name(sheet_name))
        if not entries:
            sheet.append(["Keine Daten vorhanden"])
            continue

        model = entries[0].__class__
        header = [col.name for col in model.__table__.columns]
        rows = []
        for e in entries:
            row = []
            for h in header:
                v = getattr(e, h)
                if isinstance(v, (uuid.UUID, datetime)):
                    v = str(v)
                row.append(v)
            rows.append(row)

        start_row = 10
        start_col = 2

        for col_offset, title in enumerate(header):
            col_idx = start_col + col_offset
            cell = sheet.cell(row=start_row, column=col_idx, value=title)
            cell.font = Font(bold=True)
            cell.border = table_border(
                start_row, col_idx, start_row, start_col, rows, header
            )

        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        for row_offset, row_data in enumerate(rows, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                row_idx = start_row + row_offset
                col_idx = col_index + (start_col - 1)
                cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = table_border(
                    row_idx, col_idx, start_row, start_col, rows, header
                )
                if (
                    col_index == 5
                    and isinstance(cell_value, (int, float))
                    and cell_value > 100
                ):
                    cell.fill = red_fill

    add_hyperlinked_sheet_overview(wb)
    wb.save(export_path)
    logger.success("Excel mit Analysen gespeichert unter: {}", export_path)

    # Kunden pro Jahr
    if "customer_kpis" in kpi_data and kpi_data["customer_kpis"]:
        ws = wb.create_sheet("Kunden pro Jahr")
        ws.append(["Jahr", "Anzahl"])
        stats = defaultdict(int)
        for c in kpi_data["customer_kpis"]:
            if in_range(c.registered_at):
                stats[c.registered_at.year] += 1
        for year, count in sorted(stats.items()):
            ws.append([year, count])
        chart = BarChart()
        chart.title = "Kunden pro Jahr"
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, "D2")

    # Ø Transaktionshöhe pro Monat
    if "transaction_kpis" in kpi_data and kpi_data["transaction_kpis"]:
        ws = wb.create_sheet("Ø Transaktionshöhe")
        ws.append(["Monat", "Ø Betrag"])
        monthly = defaultdict(list)
        for t in kpi_data["transaction_kpis"]:
            if in_range(t.created_at):
                key = t.created_at.strftime("%Y-%m")
                monthly[key].append(t.amount)
        for month, values in sorted(monthly.items()):
            ws.append([month, round(sum(values) / len(values), 2)])
        chart = LineChart()
        chart.title = "Ø Transaktionshöhe pro Monat"
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, "D2")

    # Bewegungstypen (Kreis)
    if "product_movement_kpis" in kpi_data and kpi_data["product_movement_kpis"]:
        ws = wb.create_sheet("Bewegung (Kreis)")
        ws.append(["Typ", "Anzahl"])
        stats = defaultdict(int)
        for m in kpi_data["product_movement_kpis"]:
            if in_range(m.created_at):
                stats[m.movement_type] += 1
        for typ, count in stats.items():
            ws.append([typ, count])
        chart = PieChart()
        chart.title = "Produktbewegung nach Typ"
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, "D2")

    # Top 5 Kunden (nach Order)
    if "order_kpis" in kpi_data and kpi_data["order_kpis"]:
        ws = wb.create_sheet("Top 5 Kunden")
        ws.append(["User", "Summe"])
        stats = defaultdict(float)
        for o in kpi_data["order_kpis"]:
            if in_range(o.created_at):
                stats[o.user_id] += o.total_price
        for user, value in sorted(stats.items(), key=lambda x: -x[1])[:5]:
            ws.append([user, value])
        chart = BarChart()
        chart.title = "Top 5 Kunden nach Bestellwert"
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, "D2")

    # Registrierung Quartale
    if "customer_kpis" in kpi_data and kpi_data["customer_kpis"]:
        ws = wb.create_sheet("Registrierungen Q")
        ws.append(["Quartal", "Anzahl"])
        stats = defaultdict(int)
        for c in kpi_data["customer_kpis"]:
            if in_range(c.registered_at):
                q = f"Q{((c.registered_at.month - 1)//3) + 1} {c.registered_at.year}"
                stats[q] += 1
        for quartal, count in sorted(stats.items()):
            ws.append([quartal, count])
        chart = BarChart()
        chart.title = "Registrierungen pro Quartal"
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, "D2")
