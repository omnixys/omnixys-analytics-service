# src/kpi/utils/export_utils.py

from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from sqlalchemy.ext.asyncio import AsyncSession
import re
import uuid
from loguru import logger

from kpi.models.entities.customer_kpi import CustomerKPI
from kpi.models.entities.transaction_kpi import TransactionKPI
from kpi.models.entities.order_kpi import OrderKPI
from kpi.models.entities.product_movement_kpi import ProductMovementKPI


def export_to_csv(kpi_data: dict[str, list], path: Path) -> None:
    import csv

    with open(path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")

        for sheet_name, entries in kpi_data.items():
            if not entries:
                continue

            # 💡 Dynamisch alle Spaltennamen aus dem SQLAlchemy-Modell
            model = entries[0].__class__
            columns = [c.name for c in model.__table__.columns]

            # ✍️ Optional: Abschnittsüberschrift in der CSV
            writer.writerow([f"=== {sheet_name.upper()} ==="])
            writer.writerow(columns)

            for entry in entries:
                writer.writerow([getattr(entry, col) for col in columns])


def _sanitize_sheet_name(name: str) -> str:
    invalid_chars = r"[:\\/*?[\]]"
    return re.sub(invalid_chars, "", name)[:31]


def _add_hyperlinked_sheet_overview(workbook: Workbook, start_row, start_col):
    """
    Creates a hyperlinked overview sheet as the first sheet in the workbook.

    Args:
        workbook: The Excel workbook to add the overview to
    """

    sheet = workbook.create_sheet("\u00dcbersicht", 0)

    # Set column widths
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 40

    # Add header with styling
    headers = ["Sheet", "Beschreibung"]
    sheet.append(headers)

    # Style header row
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )

    # Add hyperlinks to all other sheets
    for idx, worksheet in enumerate(workbook.worksheets[1:], start=2):
        # Add sheet name and description
        sheet.append(
            [worksheet.title, f"Analyse von {worksheet.title.lower()}"]
        )

        # Create hyperlink in first column
        hyperlink_cell = sheet.cell(row=idx, column=1)
        hyperlink_cell.hyperlink = (
            f"#'{worksheet.title}'!A1"  # Added quotes for sheet names with spaces
        )
        hyperlink_cell.font = Font(color="0563C1", underline="single")
        # Style all cells in the row
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=idx, column=col)
            cell.border = Border(
                left=Side(border_style="thin", color="D9D9D9"),
                right=Side(border_style="thin", color="D9D9D9"),
                top=Side(border_style="thin", color="D9D9D9"),
                bottom=Side(border_style="thin", color="D9D9D9"),
            )
    # Freeze header row
    sheet.freeze_panes = "A2"

    # Add a filter to the header row
    sheet.auto_filter.ref = f"A1:B{len(workbook.worksheets)}"


def in_range(obj_date, start_date, end_date):
    if not start_date and not end_date:
        return True
    if start_date and obj_date < start_date:
        return False
    if end_date and obj_date > end_date:
        return False
    return True


def table_border(
    row_idx: int,
    col_idx: int,
    start_row: int,
    start_col: int,
    rows: list[list],
    headers: list[str],
) -> Border:
    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    is_top = row_idx == start_row
    is_bottom = row_idx in {start_row, start_row + len(rows)}
    is_left = col_idx == start_col
    is_right = col_idx == start_col + len(headers) - 1
    return Border(
        top=thick if is_top else thin,
        bottom=thick if is_bottom else thin,
        left=thick if is_left else thin,
        right=thick if is_right else thin,
    )


def export_to_excel(
    kpi_data: dict[str, list], export_path: Path, start_date=None, end_date=None
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, entries in kpi_data.items():
        sheet = wb.create_sheet(_sanitize_sheet_name(sheet_name))

        # Set column widths
        sheet.column_dimensions["A"].width = 40
        sheet.column_dimensions["B"].width = 40
        sheet.column_dimensions["C"].width = 40
        sheet.column_dimensions["D"].width = 40

        if not entries:
            sheet.append(["Keine Daten vorhanden"])
            continue

        model = entries[0].__class__
        headers = [col.name for col in model.__table__.columns]
        rows = []
        for e in entries:
            row = []
            for h in headers:
                v = getattr(e, h)
                if isinstance(v, (uuid.UUID, datetime)):
                    v = str(v)
                row.append(v)
            rows.append(row)

        start_row = 1
        start_col = 2

        # Kopfzeile in Zeile 10 (B10)
        for col_offset, title in enumerate(headers):
            col_idx = start_col + col_offset
            cell = sheet.cell(row=start_row, column=col_idx, value=title)
            cell.font = Font(bold=True)
            cell.border = table_border(
                start_row, col_idx, start_row, start_col, rows, headers
            )

        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        for row_offset, row_data in enumerate(rows, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                row_idx = start_row + row_offset
                col_idx = col_index + (start_col - 1)
                cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = table_border(
                    row_idx, col_idx, start_row, start_col, rows, headers
                )
                if (
                    col_index == 5
                    and isinstance(cell_value, (int, float))
                    and float(cell_value) > 100
                ):
                    cell.fill = red_fill

        # Freeze header row
        sheet.freeze_panes = "A2"

        # Add a filter to the header row
        sheet.auto_filter.ref = f"A1:B{len(wb.worksheets)}"

    _kunden_pro_jahr(kpi_data, wb, start_date, end_date)
    _transaktion_pro_monat(kpi_data, wb, start_date, end_date)
    _bewegung(kpi_data, wb, start_date, end_date)
    _top_5_kunden(kpi_data, wb, start_date, end_date)
    _registrierungen_pro_quartal(kpi_data, wb, start_date, end_date)

    for ws in wb.worksheets[1:]:
        logger.debug("export_to_excel title={}", ws.title)

    _add_hyperlinked_sheet_overview(wb, start_row, start_col)
    wb.save(export_path)
    logger.success("Excel mit Analysen gespeichert unter: {}", export_path)


def _kunden_pro_jahr(kpi_data: dict[str, list], wb: Workbook, start_date, end_date):
    # Kunden pro Jahr (Säule)

    if kpi_data.get("customer_kpis"):
        sheet = wb.create_sheet("Kunden pro Jahr")
        sheet.append(["Jahr", "Anzahl"])
        stats = defaultdict(int)
        for c in kpi_data["customer_kpis"]:
            if in_range(c.registered_at, start_date, end_date):
                stats[c.registered_at.year] += 1
        for year, count in sorted(stats.items()):
            sheet.append([year, count])
        chart = BarChart()
        chart.title = "Kunden pro Jahr"
        chart.add_data(
            Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
        )
        sheet.add_chart(chart, "D2")


def _transaktion_pro_monat(kpi_data, wb, start_date, end_date):
    # Transaktionen: Ø pro Monat (Linie)
    if kpi_data.get("transaction_kpis"):
        ws = wb.create_sheet("\u00d8 Transaktionsh\u00f6he")
        ws.append(["Monat", "\u00d8 Betrag"])
        monthly = defaultdict(list)
        for t in kpi_data["transaction_kpis"]:
            if in_range(t.created_at, start_date, end_date):
                key = t.created_at.strftime("%Y-%m")
                monthly[key].append(t.amount)
        for k, lst in sorted(monthly.items()):
            avg = sum(lst) / len(lst)
            ws.append([k, avg])
        chart = LineChart()
        chart.title = "Durchschnittliche Transaktionen je Monat"
        chart.add_data(
            Reference(ws, min_col=2, min_row=1, max_row=ws.max_row),
            titles_from_data=True,
        )
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, "D2")


def _bewegung(kpi_data, wb, start_date, end_date):
    # Bewegungen: Kuchendiagramm
    if kpi_data.get("product_movement_kpis"):
        ws = wb.create_sheet("Bewegung (Kreis)")
        ws.append(["Typ", "Anzahl"])
        stat = defaultdict(int)
        for m in kpi_data["product_movement_kpis"]:
            if in_range(m.created_at, start_date, end_date):
                stat[m.movement_type] += 1
        for t, c in stat.items():
            ws.append([t, c])
        chart = PieChart()
        chart.title = "Verk\u00e4ufe vs. K\u00e4ufe"
        chart.add_data(
            Reference(ws, min_col=2, min_row=1, max_row=ws.max_row),
            titles_from_data=True,
        )
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, "D2")


def _top_5_kunden(kpi_data, wb, start_date, end_date):
    # Orders: Top 5 Kunden (Säule)
    if kpi_data.get("order_kpis"):
        ws = wb.create_sheet("Top Kunden")
        ws.append(["User", "Summe"])
        totals = defaultdict(float)
        for o in kpi_data["order_kpis"]:
            if in_range(o.created_at, start_date, end_date):
                totals[o.user_id] += o.total_price
        for uid, val in sorted(totals.items(), key=lambda x: -x[1])[:5]:
            ws.append([uid, val])
        chart = BarChart()
        chart.title = "Top 5 Kunden nach Umsatz"
        chart.add_data(
            Reference(ws, min_col=2, min_row=1, max_row=ws.max_row),
            titles_from_data=True,
        )
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, "D2")


def _registrierungen_pro_quartal(kpi_data, wb, start_date, end_date):
    # Registrierungen pro Quartal (Säule)
    if kpi_data.get("customer_kpis"):
        ws = wb.create_sheet("Registrierung Quartal")
        ws.append(["Quartal", "Anzahl"])
        qstats = defaultdict(int)
        for c in kpi_data["customer_kpis"]:
            if in_range(c.registered_at, start_date, end_date):
                q = f"Q{((c.registered_at.month-1)//3)+1} {c.registered_at.year}"
                qstats[q] += 1
        for q, val in sorted(qstats.items()):
            ws.append([q, val])
        chart = BarChart()
        chart.title = "Registrierungen pro Quartal"
        chart.add_data(
            Reference(ws, min_col=2, min_row=1, max_row=ws.max_row),
            titles_from_data=True,
        )
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, "D2")
